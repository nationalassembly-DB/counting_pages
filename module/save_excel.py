"""
전달받은 리스트를 이용해 파일명 페이지 수 등을 엑셀에 저장합니다
"""

import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill


def save_pages_to_excel(pages, excel_file):
    """리스트를 전달받아 파일명, 페이지 수, 경로명을 엑셀에 저장합니다"""
    if os.path.exists(excel_file):
        wb = load_workbook(excel_file)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        headers = ["연번", "파일명", "페이지 수", "경로명"]
        header_color = PatternFill(start_color='4f81bd',
                                   end_color='4f81bd', fill_type='solid')
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)
            ws.cell(row=1, column=col_idx).fill = header_color

    for i, page in enumerate(pages, start=ws.max_row):
        ws.append([i, page['파일명'], page['페이지 수'], page['경로명']])

    wb.save(excel_file)
